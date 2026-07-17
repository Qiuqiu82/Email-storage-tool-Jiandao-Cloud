import json
import re
import typing
from io import BytesIO


data = {
    "result": "",
    "result.success": False,
    "result.inference": False,
    "attachments": [],
    "variables": {}
}


TABLE_HEADER_KEYWORDS = ["货主", "作业类型", "作业日期", "订单号", "物料号", "品名"]
HEADER_ALIASES = {
    "shipper_name": ["货主"],
    "source_job_type": ["作业类型"],
    "job_date": ["作业日期"],
    "job_time": ["作业时间"],
    "job_code": ["作业编号"],
    "order_number": ["订单号"],
    "material_code": ["物料号"],
    "product_name": ["品名"],
    "batch_number": ["批号"],
    "unit": ["单位"],
    "quantity": ["数量"],
    "status": ["状态"],
    "specification": ["规格"],
    "weight": ["重量"],
}


def _new_data():
    return {
        "result": "",
        "result.success": False,
        "result.inference": False,
        "attachments": [],
        "variables": {}
    }


def _safe_get(source: dict, *keys, default=None):
    for key in keys:
        if isinstance(source, dict) and source.get(key) not in (None, ""):
            return source.get(key)
    return default


def _normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join([str(item) for item in value if item not in (None, "")]).strip()
    return str(value).strip()


def _download_xlsx(oss_url):
    import requests

    response = requests.get(oss_url, timeout=60)
    if response.status_code != 200:
        raise Exception(f"download file failed: {response.status_code}")
    return BytesIO(response.content)


def _cell_text(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y/%m/%d")
        except Exception:
            return str(value).strip()
    return str(value).strip()


def _row_text(row):
    return " ".join([_cell_text(cell) for cell in row if _cell_text(cell)])


def _extract_sheet_rows(xlsx_bytes):
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_bytes, data_only=True, read_only=True)
    worksheet = workbook.worksheets[0]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _find_table_start(rows):
    for row_index, row in enumerate(rows):
        row_text = _row_text(row)
        hit_count = sum(1 for keyword in TABLE_HEADER_KEYWORDS if keyword in row_text)
        if hit_count >= 4:
            return row_index
    return None


def _find_column_index(header_row, aliases):
    for column_index, header_cell in enumerate(header_row):
        header_text = _cell_text(header_cell)
        for alias in aliases:
            if alias in header_text:
                return column_index
    return None


def _is_valid_order(order):
    job_type = _normalize_text(order.get("source_job_type", ""))
    order_number = _normalize_text(order.get("order_number", ""))
    material_code = _normalize_text(order.get("material_code", ""))
    product_name = _normalize_text(order.get("product_name", ""))
    quantity = _normalize_text(order.get("quantity", ""))
    weight = _normalize_text(order.get("weight", ""))
    meaningful_count = sum(
        1 for value in [order_number, material_code, product_name, quantity, weight] if value
    )
    return bool(job_type and meaningful_count >= 2)


def _parse_orders(rows, start_row):
    header_row = rows[start_row]
    column_mapping = {
        field_name: _find_column_index(header_row, aliases)
        for field_name, aliases in HEADER_ALIASES.items()
    }

    orders = []
    for row in rows[start_row + 1:]:
        if not _row_text(row):
            if orders:
                break
            continue

        order = {}
        for field_name, column_index in column_mapping.items():
            if column_index is not None and column_index < len(row):
                order[field_name] = _cell_text(row[column_index])
            else:
                order[field_name] = ""

        if _is_valid_order(order):
            orders.append(order)
        elif orders:
            break

    return orders


def _find_label_value(rows, labels, min_row=0, max_row=None, start_col=0, col_offset=1):
    end_row = len(rows) if max_row is None else min(max_row, len(rows))
    for row_index in range(min_row, end_row):
        row = rows[row_index]
        for col_index in range(start_col, len(row)):
            cell_text = _cell_text(row[col_index])
            if cell_text and any(label in cell_text for label in labels):
                value_col = col_index + col_offset
                if value_col < len(row):
                    return _cell_text(row[value_col])
    return ""


def _find_section_row(rows, keyword):
    for row_index, row in enumerate(rows):
        if keyword in _row_text(row):
            return row_index
    return 0


def _parse_email_body_fields(mail_body):
    text = _normalize_text(mail_body)
    patterns = {
        "vehicle_number": r"(?:车牌|车牌号码)\s*[:：]?\s*([A-Z\u4e00-\u9fa50-9\-]+)",
        "driver_name": r"司机姓名\s*[:：]?\s*([^\s\r\n]+)",
        "driver_phone": r"司机.*?联系电话\s*[:：]?\s*([0-9\-]{7,20})",
        "id_number": r"司机.*?身份证号\s*[:：]?\s*([0-9Xx]{15,18})",
        "escort_name": r"押运员姓名\s*[:：]?\s*([^\s\r\n]+)",
        "escort_phone": r"押运员联系电话\s*[:：]?\s*([0-9\-]{7,20})",
        "escort_id_number": r"押运人身份证号\s*[:：]?\s*([0-9Xx]{15,18})",
    }
    parsed = {}
    for field_name, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            parsed[field_name] = match.group(1).strip()
    return parsed


def _build_customer_info(email_info, rows):
    start_row = _find_section_row(rows, "客户信息")
    end_row = min(start_row + 8, len(rows))
    info = {
        "customer_company_name": _find_label_value(rows, ["公司名称"], min_row=start_row, max_row=end_row),
        "customer_contact_name": _find_label_value(rows, ["联系人"], min_row=start_row, max_row=end_row),
        "customer_phone": _find_label_value(rows, ["联系电话"], min_row=start_row, max_row=end_row),
        "customer_email": _find_label_value(rows, ["E-mail"], min_row=start_row, max_row=end_row),
        "customer_address": _find_label_value(rows, ["公司地址"], min_row=start_row, max_row=end_row),
    }
    return info


def _build_driver_vehicle_info(rows, body_fields):
    start_row = _find_section_row(rows, "司机车辆信息")
    end_row = min(start_row + 10, len(rows))
    info = {
        "vehicle_number": _find_label_value(rows, ["车牌号码"], min_row=start_row, max_row=end_row, start_col=5) or body_fields.get("vehicle_number", ""),
        "trailer_number": _find_label_value(rows, ["挂车号码"], min_row=start_row, max_row=end_row, start_col=5),
        "driver_name": _find_label_value(rows, ["司机姓名"], min_row=start_row, max_row=end_row, start_col=5) or body_fields.get("driver_name", ""),
        "driver_phone": _find_label_value(rows, ["联系电话"], min_row=start_row + 3, max_row=start_row + 5, start_col=5) or body_fields.get("driver_phone", ""),
        "id_number": _find_label_value(rows, ["身份证号"], min_row=start_row + 4, max_row=start_row + 6, start_col=5) or body_fields.get("id_number", ""),
        "escort_name": _find_label_value(rows, ["押运员姓名"], min_row=start_row, max_row=end_row, start_col=5) or body_fields.get("escort_name", ""),
        "escort_phone": _find_label_value(rows, ["联系电话"], min_row=start_row + 5, max_row=start_row + 7, start_col=5) or body_fields.get("escort_phone", ""),
        "escort_id_number": _find_label_value(rows, ["身份证号"], min_row=start_row + 6, max_row=start_row + 8, start_col=5) or body_fields.get("escort_id_number", ""),
    }
    return info


def main(*args, tool_args: dict, **kwargs) -> typing.Any:
    data = _new_data()
    variables = kwargs.get("variables", {}) or {}
    email_info = variables.get("email_info", {}) or {}
    attachments = email_info.get("attachments", []) or []

    if variables.get("is_inoutbound_mail") is False or variables.get("mail_intent") == "其他":
        data["result.success"] = True
        data["result.inference"] = True
        data["result"] = "邮件意图不是出入库，已跳过附件订单解析"
        data["variables"]["source_order_info"] = []
        data["variables"]["customer_info"] = []
        data["variables"]["driver_vehicle_info"] = []
        data["variables"]["company_name"] = ""
        data["variables"]["job_type"] = ""
        data["variables"]["attachment_filename"] = ""
        data["variables"]["attachment_oss_url"] = ""
        return data

    if not attachments:
        data["result.success"] = True
        data["result.inference"] = True
        data["result"] = "no attachments"
        data["variables"]["source_order_info"] = []
        data["variables"]["customer_info"] = []
        data["variables"]["driver_vehicle_info"] = []
        return data

    first_attachment = None
    for attachment in attachments:
        if isinstance(attachment, dict) and str(attachment.get("filename", "")).lower().endswith(".xlsx"):
            first_attachment = attachment
            break

    if not first_attachment:
        raise ValueError("no xlsx attachment found")

    oss_url = first_attachment.get("oss_url", "")
    if not oss_url:
        raise ValueError("missing oss_url in attachment")

    rows = _extract_sheet_rows(_download_xlsx(oss_url))
    table_start = _find_table_start(rows)
    if table_start is None:
        raise ValueError("cannot find order table header")

    body_fields = _parse_email_body_fields(email_info.get("mail_body", ""))
    orders = _parse_orders(rows, table_start)
    customer_info = _build_customer_info(email_info, rows)
    driver_vehicle_info = _build_driver_vehicle_info(rows, body_fields)

    company_name = customer_info.get("customer_company_name", "")
    job_type = _safe_get(email_info, "job_type", default="")
    if not job_type and orders:
        job_type = _safe_get(orders[0], "source_job_type", default="")

    data["variables"]["source_order_info"] = orders
    data["variables"]["customer_info"] = [customer_info]
    data["variables"]["driver_vehicle_info"] = [driver_vehicle_info]
    data["variables"]["company_name"] = company_name
    data["variables"]["job_type"] = job_type
    data["variables"]["attachment_filename"] = first_attachment.get("filename", "")
    data["variables"]["attachment_oss_url"] = oss_url
    data["result.success"] = True
    data["result.inference"] = True
    data["result"] = json.dumps(
        {
            "company_name": company_name,
            "job_type": job_type,
            "order_count": len(orders),
            "driver_vehicle_info": driver_vehicle_info,
        },
        ensure_ascii=False
    )
    return data
